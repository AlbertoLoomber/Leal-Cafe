from flask import render_template, request, jsonify, session, send_file
import logging
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import contabilidad_bp
from auth import login_required
from .database import (
    obtener_gastos_mes,
    insertar_gasto,
    actualizar_gasto,
    eliminar_gasto,
    obtener_gasto_por_id,
    obtener_metricas_gastos,
    obtener_estado_resultados
)

logger = logging.getLogger(__name__)


@contabilidad_bp.route('/')
@login_required
def index():
    """Página principal de Gestión de Gastos"""
    return render_template('contabilidad/index.html')


@contabilidad_bp.route('/resumen')
@login_required
def resumen():
    """Resumen financiero - Estado de Resultados"""
    try:
        # Obtener mes y año de parámetros o usar mes actual
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)

        # Obtener datos del estado de resultados
        datos = obtener_estado_resultados(mes, anio)

        # Nombres de meses para el filtro
        meses_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }

        return render_template(
            'contabilidad/resumen.html',
            datos=datos,
            mes=mes,
            anio=anio,
            meses_nombres=meses_nombres
        )

    except Exception as e:
        logger.error(f"Error en resumen financiero: {str(e)}")
        import traceback
        traceback.print_exc()
        # Retornar con datos vacíos en caso de error
        return render_template(
            'contabilidad/resumen.html',
            datos={
                'ingresos': 0,
                'costo_ventas': 0,
                'utilidad_bruta': 0,
                'gastos_fijos': {'total': 0, 'desglose': []},
                'gastos_variables': {'total': 0, 'desglose': []},
                'utilidad_neta': 0,
                'margen_neto': 0,
                'margen_bruto': 0
            },
            mes=datetime.now().month,
            anio=datetime.now().year,
            meses_nombres={
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
            }
        )


# ==================== API ENDPOINTS ====================

@contabilidad_bp.route('/api/gastos', methods=['GET'])
@login_required
def api_obtener_gastos():
    """
    API: Obtener gastos de un mes
    Query params: mes, anio, sucursal (opcional)
    """
    try:
        mes = request.args.get('mes', type=int)
        anio = request.args.get('anio', type=int)
        sucursal = request.args.get('sucursal', None)

        if not mes or not anio:
            return jsonify({
                'success': False,
                'error': 'Se requieren parámetros mes y anio'
            }), 400

        datos = obtener_gastos_mes(mes, anio, sucursal)

        # Convertir fechas a string para JSON
        for gasto in datos['gastos']:
            if isinstance(gasto['fecha'], date):
                gasto['fecha'] = gasto['fecha'].strftime('%Y-%m-%d')
            if gasto.get('fecha_registro'):
                gasto['fecha_registro'] = gasto['fecha_registro'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({
            'success': True,
            'data': datos
        })

    except Exception as e:
        logger.error(f"Error en API obtener gastos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@contabilidad_bp.route('/api/gastos', methods=['POST'])
@login_required
def api_crear_gasto():
    """
    API: Crear nuevo gasto
    Body JSON: {fecha, sucursal, tipo_gasto, categoria, descripcion,
                forma_pago, monto, facturado, comentarios}
    """
    try:
        data = request.get_json()

        # Log para debugging
        logger.info(f"Datos recibidos para crear gasto: {data}")

        # Validar que se recibieron datos
        if not data:
            return jsonify({
                'success': False,
                'error': 'No se recibieron datos'
            }), 400

        # Validar campos requeridos
        campos_requeridos = ['fecha', 'sucursal', 'tipo_gasto', 'categoria', 'forma_pago', 'monto', 'descripcion']
        campos_faltantes = []

        for campo in campos_requeridos:
            if campo not in data or data[campo] == '' or data[campo] is None:
                campos_faltantes.append(campo)

        if campos_faltantes:
            return jsonify({
                'success': False,
                'error': f'Campos requeridos faltantes o vacíos: {", ".join(campos_faltantes)}'
            }), 400

        # Obtener usuario de sesión
        usuario_id = session.get('user_id')
        if not usuario_id:
            return jsonify({
                'success': False,
                'error': 'Usuario no autenticado'
            }), 401

        # Convertir facturado a boolean
        facturado_str = data.get('facturado', 'No')
        facturado = facturado_str == 'Sí'

        logger.info(f"Facturado recibido: '{facturado_str}' -> convertido a boolean: {facturado}")

        # Validar y convertir monto
        try:
            monto = float(data['monto'])
            if monto <= 0:
                return jsonify({
                    'success': False,
                    'error': 'El monto debe ser mayor a 0'
                }), 400
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'El monto debe ser un número válido'
            }), 400

        # Insertar gasto
        gasto_id = insertar_gasto(
            fecha=data['fecha'],
            sucursal=data['sucursal'],
            tipo_gasto=data['tipo_gasto'],
            categoria=data['categoria'],
            descripcion=data.get('descripcion', ''),
            forma_pago=data['forma_pago'],
            monto=monto,
            facturado=facturado,
            comentarios=data.get('comentarios', ''),
            usuario_id=usuario_id
        )

        facturado_texto = 'Sí' if facturado else 'No'
        logger.info(f"Gasto creado exitosamente con ID: {gasto_id} (Facturado: {facturado_texto})")

        return jsonify({
            'success': True,
            'message': f'Gasto registrado exitosamente - Facturado: {facturado_texto}',
            'gasto_id': gasto_id
        })

    except Exception as e:
        logger.error(f"Error en API crear gasto: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500


@contabilidad_bp.route('/api/gastos/<int:gasto_id>', methods=['PUT'])
@login_required
def api_actualizar_gasto(gasto_id):
    """
    API: Actualizar gasto existente
    Body JSON: {fecha, sucursal, tipo_gasto, categoria, descripcion,
                forma_pago, monto, facturado, comentarios}
    """
    try:
        data = request.get_json()

        # Validar campos requeridos
        campos_requeridos = ['fecha', 'sucursal', 'tipo_gasto', 'categoria', 'forma_pago', 'monto']
        for campo in campos_requeridos:
            if campo not in data or data[campo] == '':
                return jsonify({
                    'success': False,
                    'error': f'El campo {campo} es requerido'
                }), 400

        # Convertir facturado a boolean
        facturado = data.get('facturado', 'No') == 'Sí'

        # Actualizar gasto
        success = actualizar_gasto(
            gasto_id=gasto_id,
            fecha=data['fecha'],
            sucursal=data['sucursal'],
            tipo_gasto=data['tipo_gasto'],
            categoria=data['categoria'],
            descripcion=data.get('descripcion', ''),
            forma_pago=data['forma_pago'],
            monto=float(data['monto']),
            facturado=facturado,
            comentarios=data.get('comentarios', '')
        )

        if success:
            return jsonify({
                'success': True,
                'message': 'Gasto actualizado exitosamente'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Gasto no encontrado'
            }), 404

    except Exception as e:
        logger.error(f"Error en API actualizar gasto {gasto_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@contabilidad_bp.route('/api/gastos/<int:gasto_id>', methods=['DELETE'])
@login_required
def api_eliminar_gasto(gasto_id):
    """
    API: Eliminar gasto
    """
    try:
        success = eliminar_gasto(gasto_id)

        if success:
            return jsonify({
                'success': True,
                'message': 'Gasto eliminado exitosamente'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Gasto no encontrado'
            }), 404

    except Exception as e:
        logger.error(f"Error en API eliminar gasto {gasto_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@contabilidad_bp.route('/api/gastos/<int:gasto_id>', methods=['GET'])
@login_required
def api_obtener_gasto(gasto_id):
    """
    API: Obtener un gasto específico por ID
    """
    try:
        gasto = obtener_gasto_por_id(gasto_id)

        if gasto:
            # Convertir fechas a string
            if isinstance(gasto['fecha'], date):
                gasto['fecha'] = gasto['fecha'].strftime('%Y-%m-%d')
            if gasto.get('fecha_registro'):
                gasto['fecha_registro'] = gasto['fecha_registro'].strftime('%Y-%m-%d %H:%M:%S')

            return jsonify({
                'success': True,
                'data': gasto
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Gasto no encontrado'
            }), 404

    except Exception as e:
        logger.error(f"Error en API obtener gasto {gasto_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@contabilidad_bp.route('/api/metricas', methods=['GET'])
@login_required
def api_obtener_metricas():
    """
    API: Obtener métricas de gastos
    Query params: mes, anio
    """
    try:
        mes = request.args.get('mes', type=int)
        anio = request.args.get('anio', type=int)

        if not mes or not anio:
            return jsonify({
                'success': False,
                'error': 'Se requieren parámetros mes y anio'
            }), 400

        metricas = obtener_metricas_gastos(mes, anio)

        return jsonify({
            'success': True,
            'data': metricas
        })

    except Exception as e:
        logger.error(f"Error en API métricas: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@contabilidad_bp.route('/exportar-excel', methods=['GET'])
@login_required
def exportar_excel():
    """
    Exportar gastos a Excel
    Query params: mes, anio
    """
    try:
        mes = request.args.get('mes', type=int, default=1)
        anio = request.args.get('anio', type=int, default=2026)

        # Obtener datos
        datos = obtener_gastos_mes(mes, anio)
        gastos = datos['gastos']

        # Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Gastos {mes}-{anio}"

        # Estilos
        header_fill = PatternFill(start_color="8B6914", end_color="8B6914", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezados
        headers = [
            'Fecha', 'Sucursal', 'Tipo de Gasto', 'Categoría', 'Descripción',
            'Forma de Pago', 'Monto', '%', '¿Facturado?', 'Comentarios'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos
        meses_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
            7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }

        for row, gasto in enumerate(gastos, 2):
            # Convertir fecha
            if isinstance(gasto['fecha'], str):
                fecha_str = datetime.strptime(gasto['fecha'], '%Y-%m-%d').strftime('%d/%m/%Y')
            else:
                fecha_str = gasto['fecha'].strftime('%d/%m/%Y')

            ws.cell(row=row, column=1).value = fecha_str
            ws.cell(row=row, column=2).value = gasto['sucursal']
            ws.cell(row=row, column=3).value = gasto['tipo_gasto']
            ws.cell(row=row, column=4).value = gasto['categoria']
            ws.cell(row=row, column=5).value = gasto['descripcion'] or ''
            ws.cell(row=row, column=6).value = gasto['forma_pago']
            ws.cell(row=row, column=7).value = float(gasto['monto'])
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
            ws.cell(row=row, column=8).value = float(gasto['porcentaje']) / 100
            ws.cell(row=row, column=8).number_format = '0.0%'
            ws.cell(row=row, column=9).value = 'Sí' if gasto['facturado'] else 'No'
            ws.cell(row=row, column=10).value = gasto['comentarios'] or ''

            # Aplicar bordes
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = border

        # Fila de totales
        total_row = len(gastos) + 2
        ws.cell(row=total_row, column=6).value = 'TOTAL:'
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')

        ws.cell(row=total_row, column=7).value = float(datos['total_mes'])
        ws.cell(row=total_row, column=7).number_format = '$#,##0.00'
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=7).fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        ws.cell(row=total_row, column=8).value = 1.0
        ws.cell(row=total_row, column=8).number_format = '0.0%'
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=8).fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12  # Fecha
        ws.column_dimensions['B'].width = 12  # Sucursal
        ws.column_dimensions['C'].width = 14  # Tipo
        ws.column_dimensions['D'].width = 16  # Categoría
        ws.column_dimensions['E'].width = 30  # Descripción
        ws.column_dimensions['F'].width = 14  # Forma de Pago
        ws.column_dimensions['G'].width = 12  # Monto
        ws.column_dimensions['H'].width = 8   # %
        ws.column_dimensions['I'].width = 12  # Facturado
        ws.column_dimensions['J'].width = 30  # Comentarios

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nombre del archivo
        nombre_mes = meses_nombres.get(mes, mes)
        filename = f"Gastos_{nombre_mes}_{anio}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Error exportando a Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
